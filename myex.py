import openpyxl as xl

exf = xl.load_workbook('c:\\dd\\ss.xlsx') # open - 파일 연결, load - 

aws = exf.active

for i in aws.rows:
    index = i[0].row # 인덱스 번호와 상관이 없다...
    name = i[0].value
    kor = i[1].value
    eng = i[2].value
    math = i[3].value

    total = kor + eng + math
    avg = total / 3

    aws.cell(row = index, column = 5).value = total # cell 함수
    aws.cell(row = index, column = 6).value = avg

    print('{} {} {} {} {} {:.2f}'.format(name, kor, eng, math, total, avg))

exf.save('c:\\dd\\outss.xlsx') # 새로운 파일로 지정한 디렉토리에 저장
exf.close()
